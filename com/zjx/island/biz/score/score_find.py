import openpyxl


class ScoreFind:
    def __init__(self, workbook_name):
        self.worksheet = None
        self.workbook = None
        self.workbook_name = workbook_name

    def open_read_default_sheet(self):
        self.workbook = openpyxl.load_workbook(self.workbook_name)
        self.worksheet = self.workbook.active

    def get_all_rows_without_first(self):
        rows_without_first = []
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            rows_without_first.append(row)
        self.workbook.close
        return rows_without_first

    def get_all_rows(self):
        rows_without_first = []
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            rows_without_first.append(row)
        self.workbook.close
        return rows_without_first


if __name__ == '__main__':
    wb1 = openpyxl.load_workbook('./find/w1.xlsx')
    sheet = wb1['Sheet1']  # 假设工作簿中有一个名为"Sheet1"的sheet
    origin_rows = []
    target_rows = []
    for row in sheet.iter_rows(values_only=True):
        origin_rows.append(row[1:3])
        target_rows.append((row[-1],None))

    print(origin_rows)
    print(target_rows)
    print(len(origin_rows))
    print(len(target_rows))
    for target in target_rows:
        for origin in origin_rows:
            if target[0] is not None:
                if origin[0] in target[0]:
                    print(origin[1])


