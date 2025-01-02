import os

import openpyxl
from com.zjx.island.biz.score.yaml_util import get_all_from_columns, load_yaml

RESULT_WORKBOOK_NAME = '自动合分结果.xlsx'


class ScoreStatistics:
    def __init__(self, workbook_name):
        self.worksheet = None
        self.workbook = None
        self.workbook_name = workbook_name

    def open_read_default_sheet(self):
        self.workbook = openpyxl.load_workbook(self.workbook_name)
        self.worksheet = workbook.active

    def get_all_rows_without_first(self):
        rows_without_first = []
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            rows_without_first.append(row)
        self.workbook.close
        return rows_without_first



def get_all_rows_without_first(path):
    # 打开一个已存在的工作簿
    workbook = openpyxl.load_workbook(path)

    # 获取默认的工作表
    sheet = workbook.active

    origin_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        origin_rows.append(row)

    # 关闭工作簿
    workbook.close()
    return origin_rows

def get_all_rows_without_first_from_all_sheet(path):
    # 打开一个已存在的工作簿
    workbook = openpyxl.load_workbook(path)

    all_rows_data = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        origin_rows = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            origin_rows.append(row)
        all_rows_data.append(origin_rows)

    # 关闭工作簿
    workbook.close()
    return all_rows_data



def select_need_column(origin_rows, column_names):
    # 指定要选择的下标位置
    selected_indices = []
    selected_rows = []

    head_line = origin_rows[0]
    # 根据列名找出要筛选的下标
    for cn in column_names:
        print(cn)
        print(head_line)
        index = head_line.index(cn)
        selected_indices.append(index)
    for row in origin_rows:
        selected_rows.append(tuple(row[i] for i in selected_indices))
    # 排除学号为空以及为-的
    selected_rows_without_none = [r for r in selected_rows if r[0] is not None and r[0] != '-']
    return selected_rows_without_none


def convert_to_target_column(origin_rows, column_mapping):
    result_rows = []
    # 添加表头
    result_rows.append(tuple([m['target'] for m in column_mapping]))
    head_line = origin_rows[0]

    # 找到来源列的index
    for m in column_mapping:
        target_column_name = m['target']
        from_column_name_list = m['from']
        m['from_index'] = []
        for c in from_column_name_list:
            m['from_index'].append(head_line.index(c))

    for r in origin_rows[1:]:
        # 每个结果行作为list存储
        single_result_row_list = []
        for m in column_mapping:
            value = 0
            if len(m['from_index']) > 1:
                # 处理多个下标的情况，将对应索引的列值相加
                for i in m['from_index']:
                    if r[i] == '-':
                        value = '-'
                        break
                    else:
                        value += r[i]
            else:
                value = r[m['from_index'][0]]
            single_result_row_list.append(value)
        result_rows.append(tuple(single_result_row_list))
    return result_rows


def write_data_to_new_sheet(path):
    # 创建一个新的工作簿
    workbook = openpyxl.load_workbook(path)

    # 创建一个新的工作表（sheet）
    new_sheet = workbook.create_sheet(title='NewSheet')

    # 示例数据
    data = [
        ['Name', 'Age', 'City'],
        ['John', 25, 'New York'],
        ['Alice', 30, 'London'],
        ['Bob', 22, 'Tokyo']
    ]

    # 将数据写入新的工作表
    for row_data in data:
        new_sheet.append(row_data)

    # 保存工作簿到文件
    workbook.save('example.xlsx')


# 从第n行开始获取数据
def get_rows_from_n_start(path, n=1):
    # 打开一个已存在的工作簿
    workbook = openpyxl.load_workbook(path)

    # 获取默认的工作表
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(min_row=n, values_only=True):
        rows.append(row)

    # 关闭工作簿
    workbook.close()
    return rows

# 获取一个工作表所有sheet页的数据，按sheet页排成一个嵌套列表
def get_all_rows_without_first_from_all_sheet(path):
    # 打开一个已存在的工作簿
    workbook = openpyxl.load_workbook(path)

    all_rows_data = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        origin_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            origin_rows.append(row)
        all_rows_data.append(origin_rows)

    # 关闭工作簿
    workbook.close()
    return all_rows_data


# original_tuple = (1, 2, 3, 4, 5, 6)
#
# # 指定要选择的下标位置
# selected_indices = (0, 2, 4)
#
# # 使用切片生成新的元组
# new_tuple = tuple(original_tuple[i] for i in selected_indices)
#
# # 打印新的元组
# print(new_tuple)
#
# column_names = ['姓名', '班内学号', '单选题', '31(1)', '31(2)', '31(3)', '31(4)', '32(1)', '32(2)', '32(3)',
#                 '33(1)', '33(2)', '33(3)', '33(4)', '34(1)', '34(2)', '34(3)', '34(4)']
# # 按列选出需要的列
# selected_columns = (select_need_column(origin_rows, column_names))
#
#
# def sort_by_column(origin_rows, column_name):
#     column_index = origin_rows[0].index(column_name)
#     sorted_list = sorted(origin_rows, key=lambda x: x[column_index])
#     return sorted_list
#
#
# # 按指定列排序，比如学号
# print(sort_by_column(origin_rows, '班内学号'))


if __name__ == '__main__':

    folder_path = './待处理分数统计表'

    # 使用 os.listdir() 获取文件夹下的所有文件和子文件夹的名称
    files_and_folders = os.listdir(folder_path)

    # 过滤出只是文件的项
    files = [f for f in files_and_folders if os.path.isfile(os.path.join(folder_path, f))]
    files = [f for f in files if f.endswith('.xlsx')]
    files = sorted(files)
    for file in files:
        print(file)

    # 创建保存结果的工作表
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # 打印所有文件的名称
    for file in files:
        # 首先去掉顶部标题
        origin_rows = get_all_rows_without_first(os.path.join(folder_path, file))
        selected_rows = select_need_column(origin_rows, get_all_from_columns())
        result_rows = convert_to_target_column(selected_rows, load_yaml())
        # 按学号排序
        sorted_list = sorted(result_rows, key=lambda x: x[0])
        sorted_list.insert(0, sorted_list.pop())

        for r in sorted_list:
            print(r)

        new_sheet = workbook.create_sheet(title=str(file).split('.')[0].split('-')[1])

        # 将数据写入新的工作表
        for row_data in sorted_list:
            new_sheet.append(row_data)

    workbook.save(RESULT_WORKBOOK_NAME)

# if __name__ == '__main__':
#
#     file_path = './待处理分数统计表/2313_44中179.xlsx'
#
#
#     # 创建保存结果的工作表
#     workbook = openpyxl.Workbook()
#     workbook.remove(workbook.active)
#
#     all_rows_in_sheets = get_all_rows_without_first_from_all_sheet(file_path)
#
#
#     class_num = 1
#     for origin_rows in all_rows_in_sheets:
#         selected_rows = select_need_column(origin_rows, get_all_from_columns())
#         result_rows = convert_to_target_column(selected_rows, load_yaml())
#         # 按学号排序
#         sorted_list = sorted(result_rows, key=lambda x: x[0])
#         sorted_list.insert(0, sorted_list.pop())
#
#         for r in sorted_list:
#             print(r)
#
#         new_sheet = workbook.create_sheet(title=f'{class_num}班')
#
#         # 将数据写入新的工作表
#         for row_data in sorted_list:
#             new_sheet.append(row_data)
#         class_num += 1
#
#     workbook.save(RESULT_WORKBOOK_NAME)
