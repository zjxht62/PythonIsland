import logging
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import shutil
from tkinter.scrolledtext import ScrolledText

import openpyxl
from enum import Enum
from pathlib import Path

from com.zjx.island.biz.score.path_util import get_config_path, create_results_folder
from com.zjx.island.biz.score.yaml_util import load_yaml
from com.zjx.island.biz.score.score_statistics import get_all_from_columns, get_all_rows_without_first_from_all_sheet, \
    get_all_rows_without_first, select_need_column, convert_to_target_column, get_rows_from_n_start

output_dir = create_results_folder('output_files')
selected_template_path = None
selected_subject = None
selected_score_files = []
auto_result_file_name = ''


class PrintRedirector:
    """重定向标准输出到 Tkinter 的 Text 控件"""

    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.config(state=tk.NORMAL)  # 解锁控件，允许写入
        self.text_widget.insert(tk.END, message)  # 插入内容
        self.text_widget.config(state=tk.DISABLED)  # 禁止编辑
        self.text_widget.see(tk.END)  # 自动滚动到最新内容

    def flush(self):
        pass  # 必须实现 flush 方法以符合标准输出的要求


class Subject(Enum):
    YUWEN = 1
    SHUXUE = 2
    YINGYU = 3
    WULI = 4
    HUAXUE = 5
    SHENGWU = 6
    LISHI = 7
    DILI = 8
    ZHENGZHI = 9


def select_score_files():
    files = filedialog.askopenfilenames(title="选择多个文件")
    if files:
        selected_score_files.clear()
        selected_score_files.extend(files)
        update_file_list_display()
    else:
        messagebox.showwarning("警告", "请至少选择一个文件！")


def select_template_file():
    global selected_template_path
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename()
    if file_path:
        selected_template_path = Path(file_path)
        # 如果路径太长，只显示末尾部分并添加省略号
        if len(file_path) > 50:  # 假设最大显示宽度为 50 字符
            display_text = f"...{file_path[-50:]}"
        else:
            display_text = file_path
        template_file_label.config(text=display_text)
    else:
        # 如果未选择文件，显示默认提示
        template_file_label.config(text="未选择文件")
        selected_template_path = None


def update_file_list_display():
    score_file_list_display.delete(1.0, tk.END)
    for file in selected_score_files:
        score_file_list_display.insert(tk.END, file + "\n")


def generate_result_files():
    global auto_result_file_name
    if not selected_subject:
        messagebox.showwarning("警告", "没有选择科目！")
        return

    if not selected_score_files:
        messagebox.showwarning("警告", "没有选择文件！")
        return

    auto_result_file_name = f'自动合分结果-{selected_subject.name}.xlsx'


    # 创建输出文件目录
    # output_dir = create_results_folder('output_files')
    result_file_path = output_dir / auto_result_file_name

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    print(f'开始生成自动合分结果文件...')

    for file in selected_score_files:
        file_path = Path(file)
        file_name_without_suffix = file_path.stem
        print(f'当前读取的文件：{file_name_without_suffix}')

        # 首先去掉顶部标题
        # origin_rows = get_all_rows_without_first(file)
        # 高一上学期期末，无表头
        origin_rows = get_rows_from_n_start(file)
        selected_rows = select_need_column(origin_rows, get_all_from_columns(selected_subject))
        result_rows = convert_to_target_column(selected_rows, load_yaml(selected_subject))
        # 按学号排序
        sorted_list = sorted(result_rows, key=lambda x: x[0])
        sorted_list.insert(0, sorted_list.pop())

        for r in sorted_list:
            print(f'处理后的行：{r}')


        # new_sheet = workbook.create_sheet(title=str(file).split('.')[0].split('-')[1])
        # 选择文件名作为sheet页名称
        new_sheet = workbook.create_sheet(title=file_name_without_suffix)

        # 将数据写入新的工作表
        for row_data in sorted_list:
            new_sheet.append(row_data)

    workbook.save(result_file_path)
    print(f"自动合分结果文件生成成功：保存在{result_file_path.parent}目录下")
    messagebox.showinfo("成功", f"生成了新文件：{result_file_path.name}，保存在{result_file_path.parent}目录下")


def generate_copy_files():
    if not selected_subject:
        messagebox.showwarning("警告", "没有选择科目！")
        return

    if not selected_template_path or not selected_template_path.exists():
        messagebox.showerror("错误", "模板文件不存在！")
        return

    auto_to_copy_file_name = f'待复制结果-{selected_subject.name}.xlsx'
    # 读取模板文件
    # template_file = get_config_path('templates', copy_template_file_name)  # 假设模板文件为 template.txt
    template_file = selected_template_path  # 假设模板文件为 template.txt


    # 结果文件路径
    # auto_result_file = f'自动合分结果-{selected_subject.name}.xlsx'
    # output_dir = get_executable_dir() / 'output_files'
    # output_dir = create_results_folder('output_files')
    auto_result_file_path = output_dir / auto_result_file_name
    if not auto_result_file_path.exists():
        messagebox.showerror("错误", "无法找到自动合分结果文件，请先生成！")
        return
    auto_to_copy_file_path = output_dir / auto_to_copy_file_name

    print('开始生成可复制文件...')

    # 按照粘贴模板生成结果
    head_rows = get_rows_from_n_start(template_file)

    output_rows = []

    all_auto_result_rows = []
    for r in get_all_rows_without_first_from_all_sheet(auto_result_file_path):
        all_auto_result_rows += r


    for m in head_rows:
        for r in all_auto_result_rows:
            if m[1] == r[1]:
                output_rows.append(m + r)
                break
        else:
            print(f"未找到学生{m[1]}")
            output_rows.append(m + ('-',))

    print(len(output_rows))
    for r in output_rows:
        print(r)





    # 创建保存结果的工作表
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    new_sheet = workbook.create_sheet('待复制')
    # for r in output_rows:
    #     new_sheet.append(r)
    # 找出最长行的长度
    max_len = max(len(row) for row in output_rows)
    # 补全数据
    for row in output_rows:
        row = list(row)
        row += [None] * (max_len - len(row))  # 填充 None
        new_sheet.append(row)

    workbook.save(auto_to_copy_file_path)
    print(f"可复制文件生成成功：保存在{auto_to_copy_file_path.parent}目录下")

    messagebox.showinfo("成功",
                        f"生成了新文件：{auto_to_copy_file_path.name}，保存在{auto_to_copy_file_path.parent}目录下")


def on_combobox_select(event):
    global selected_subject
    """当下拉框选择项改变时触发的事件"""
    selected_value = combobox.get()  # 获取当前选中的值
    if selected_value == '语文':
        selected_subject = Subject.YUWEN
    elif selected_value == '数学':
        selected_subject = Subject.SHUXUE
    elif selected_value == '英语':
        selected_subject = Subject.YINGYU
    elif selected_value == '物理':
        selected_subject = Subject.WULI
    elif selected_value == '化学':
        selected_subject = Subject.HUAXUE
    elif selected_value == '生物':
        selected_subject = Subject.SHENGWU
    elif selected_value == '历史':
        selected_subject = Subject.LISHI
    elif selected_value == '地理':
        selected_subject = Subject.DILI
    elif selected_value == '政治':
        selected_subject = Subject.ZHENGZHI
    # logging.info(f"选中的值: {selected_subject}")
    print(f"选中的值：{selected_subject}")


def main():
    global score_file_list_display, selected_score_files, combobox, selected_subject, template_file_label, \
        selected_template_path
    # output_dir = create_results_folder('output_files')

    root = tk.Tk()
    root.title("合分小助手")
    root.geometry("800x600")

    # 配置列权重以支持居中对齐
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)

    # 创建下拉选择框（Combobox）
    options = ["语文", "数学", "英语", "物理", "化学", "生物", "历史", "地理", "生物"]
    combobox = ttk.Combobox(root, values=options)
    combobox.set("请选择学科")  # 设置默认提示文字
    combobox.bind("<<ComboboxSelected>>", on_combobox_select)
    combobox.grid(row=0, column=0, columnspan=2, padx=20, pady=10, sticky=tk.NSEW)  # 跨两列居中

    # 模板文件选择
    select_template_file_button = tk.Button(root, text="选择模板文件", command=select_template_file)
    select_template_file_button.grid(row=1, column=0, padx=20, pady=10)

    # 创建一个标签，用于显示选中文件的路径
    template_file_label = tk.Label(root, text="未选择文件", width=60, anchor="w")  # 固定宽度和左对齐
    template_file_label.grid(row=1, column=1, padx=20, pady=10, sticky=tk.W)

    # 原始分数文件选择
    select_score_files_button = tk.Button(root, text="选择原始分数文件", command=select_score_files)
    select_score_files_button.grid(row=2, column=0, padx=20, pady=10)
    score_file_list_display = tk.Text(root, height=10, width=60)
    score_file_list_display.grid(row=3, column=0, columnspan=2, padx=20, pady=10, sticky=tk.NSEW)  # 跨两列居中

    generate_button = tk.Button(root, text="生成合分结果文件", command=generate_result_files)
    generate_button.grid(row=4, column=0, padx=20, pady=10, sticky=tk.E)

    copy_button = tk.Button(root, text="生成可复制文件", command=generate_copy_files)
    copy_button.grid(row=4, column=1, padx=20, pady=10, sticky=tk.W)

    # 创建带滚动条的文本框
    log_display = ScrolledText(root, width=60, height=10)
    log_display.grid(row=5, column=0, columnspan=2, padx=20, pady=10, sticky=tk.NSEW)

    # 重定向标准输出到文本框
    # sys.stdout = PrintRedirector(log_display)
    # sys.stderr = PrintRedirector(log_display)

    root.mainloop()


if __name__ == "__main__":
    # 打开文件并重定向输出
    main()
